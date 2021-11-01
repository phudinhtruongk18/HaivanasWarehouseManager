import datetime
import openpyxl as excel
import os
from openpyxl.styles import Font, Alignment, GradientFill
from openpyxl.utils import get_column_letter


def read_warehouse(fileName, sheetname):
    # time a lot of time
    docData = excel.load_workbook(filename=fileName)
    duLieu = docData[sheetname]
    listDuLieuTemp = []
    for index, dataTemp in enumerate(duLieu.values):
        # bo ba hang đau tien
        if index < 3:
            continue
        # ket thuc neu xuat hien chu nay
        if dataTemp[0] is None or dataTemp[0] == "END_WAREHOUSE":
            break
        # lay du lieu tu 7 cot
        listDuLieuTemp.append(dataTemp[:7])
    return listDuLieuTemp


def export_warehouse(warehouse_stock, sum_in, sum_sale, sum_ava):
    ghiData = excel.Workbook()
    trangTinh = ghiData.active
    trangTinh.title = "Sheet1"

    trangTinh.append(("WAREHOUSE'S THU HUONG HAVAIANAS",))
    trangTinh.merge_cells(f'A1:F2')
    trangTinh.append(("",))
    title = trangTinh['A1']

    title.font = Font(size=19, bold=True)
    font_son = Font(size=12)
    font_father = Font(size=16, bold=True)

    fill_father = GradientFill(stop=("FFEC4F", "FFEC4F"))
    alignment = Alignment(horizontal="center", vertical="center")

    title.alignment = alignment

    trangTinh.append(("BARCODE", "ENGLISH NAME", "VIETNAMESE NAME", "IN", "SALE", "QUANTITY"))
    a = trangTinh['A3']
    b = trangTinh['B3']
    c = trangTinh['C3']
    d = trangTinh['D3']
    e = trangTinh['E3']
    f = trangTinh['F3']

    a.fill = b.fill = c.fill = d.fill = e.fill = f.fill = fill_father
    a.font = b.font = c.font = d.font = e.font = f.font = font_father

    adjust_column_width_from_col(trangTinh, 3, 1, trangTinh.max_column)

    index_cell = 4
    for index, stock in enumerate(warehouse_stock):
        trangTinh.append((stock.excel_format()))
        a = trangTinh[f'A{index_cell + index}']
        b = trangTinh[f'B{index_cell + index}']
        c = trangTinh[f'C{index_cell + index}']
        d = trangTinh[f'D{index_cell + index}']
        e = trangTinh[f'E{index_cell + index}']
        f = trangTinh[f'F{index_cell + index}']
        a.font = b.font = c.font = d.font = e.font = f.font = font_son
        a.alignment = b.alignment = c.alignment = d.alignment = e.alignment = f.alignment = alignment

    trangTinh.append(("END_WAREHOUSE", "", "", sum_in, sum_sale, sum_ava))

    linkthumuc = os.curdir
    x = datetime.datetime.now()
    date = x.strftime("%Y-%m-%d-%Hh%M")
    ghiData.save(filename=linkthumuc + "/Output/" + date + ".xlsx")


def adjust_column_width_from_col(ws, min_row, min_col, max_col):
    column_widths = []

    for i, col in \
            enumerate(
                ws.iter_cols(min_col=min_col, max_col=max_col, min_row=min_row)
            ):

        for cell in col:
            value = cell.value
            if value is not None:

                if isinstance(value, str) is False:
                    value = str(value)

                try:
                    column_widths[i] = max(column_widths[i], len(value))
                except IndexError:
                    column_widths.append(len(value))

    for i, width in enumerate(column_widths):
        col_name = get_column_letter(min_col + i)
        value = column_widths[i] + 20
        ws.column_dimensions[col_name].width = value
